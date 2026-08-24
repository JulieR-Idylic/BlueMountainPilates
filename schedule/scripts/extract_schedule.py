from pathlib import Path
import json
import re
import sys

from openpyxl import load_workbook


DEFAULT_WORKBOOK = "Studio_schedule2026.xlsx"
DEFAULT_OUTPUT = "schedule/data/schedule.json"

EXPECTED_COLUMN_COUNT = 8

EXPECTED_HEADERS = [
    "",
    "",
    "Reformer",
    "Reformer",
    "Reformer",
    "Reformer",
    "CCB",
    "Instructor",
]


class ScheduleAnomaly(Exception):
    """
    A workbook condition that the renderer does not know
    how to represent safely.

    These are Level 2 errors.
    """

    def __init__(self, code, message):
        super().__init__(message)
        self.code = code
        self.message = message


def info(code, message):
    print(f"INFO [{code}]: {message}")


def get_workbook_path() -> Path:
    if len(sys.argv) > 1:
        return Path(sys.argv[1])

    return Path(DEFAULT_WORKBOOK)


def normalize_defined_name_formula(value: str) -> str:
    value = value.strip()

    if value.startswith("="):
        value = value[1:]

    return value


def is_direct_range_reference(value: str) -> bool:
    return "!" in value


def parse_direct_range(value: str):
    match = re.match(
        r"^'?(.+?)'?!\$?([A-Z]+)\$?(\d+):\$?([A-Z]+)\$?(\d+)$",
        value,
    )

    if not match:
        raise ScheduleAnomaly(
            "INVALID_RANGE_REFERENCE",
            f"Could not parse range reference: {value}",
        )

    sheet_name = match.group(1)

    cell_range = (
        f"{match.group(2)}{match.group(3)}:"
        f"{match.group(4)}{match.group(5)}"
    )

    return sheet_name, cell_range


def resolve_defined_name(workbook, name: str, visited=None):
    if visited is None:
        visited = []

    if name in visited:
        chain = " -> ".join(visited + [name])

        raise ScheduleAnomaly(
            "CIRCULAR_DEFINED_NAME",
            f"Circular defined-name reference detected: {chain}",
        )

    visited = visited + [name]

    defined_name = workbook.defined_names.get(name)

    if defined_name is None:
        raise ScheduleAnomaly(
            "DEFINED_NAME_MISSING",
            f"Defined name not found: {name}",
        )

    reference = normalize_defined_name_formula(
        defined_name.attr_text
    )

    if is_direct_range_reference(reference):
        sheet_name, cell_range = parse_direct_range(reference)

        if sheet_name not in workbook.sheetnames:
            raise ScheduleAnomaly(
                "WORKSHEET_MISSING",
                (
                    f"Defined name {name} refers to missing "
                    f"worksheet: {sheet_name}"
                ),
            )

        return {
            "defined_name": name,
            "sheet": sheet_name,
            "range": cell_range,
        }

    resolved = resolve_defined_name(
        workbook,
        reference,
        visited,
    )

    return {
        "defined_name": name,
        "points_to": reference,
        "sheet": resolved["sheet"],
        "range": resolved["range"],
    }


def web_name_sort_key(name: str):
    if name == "Web_ThisWeek":
        return 0

    if name == "Web_NextWeek":
        return 1

    match = re.match(r"Web_WeekPlus(\d+)$", name)

    if match:
        return int(match.group(1))

    return 999


def validate_web_name(name):
    if name in {
        "Web_ThisWeek",
        "Web_NextWeek",
    }:
        return

    if re.fullmatch(
        r"Web_WeekPlus\d+",
        name,
    ):
        return

    raise ScheduleAnomaly(
        "UNKNOWN_WEB_ALIAS",
        (
            f"Found an unrecognized Web_ alias: {name}. "
            "Expected Web_ThisWeek, Web_NextWeek, "
            "or Web_WeekPlus<number>."
        ),
    )


def find_merge_for_cell(worksheet, cell):
    """
    Return merge information for a cell.

    Only the upper-left cell of a merged range is emitted.
    Other cells covered by the merge are omitted.
    """
    for merged_range in worksheet.merged_cells.ranges:
        if cell.coordinate not in merged_range:
            continue

        if (
            cell.row == merged_range.min_row
            and cell.column == merged_range.min_col
        ):
            return {
                "rowspan":
                    merged_range.max_row
                    - merged_range.min_row
                    + 1,
                "colspan":
                    merged_range.max_col
                    - merged_range.min_col
                    + 1,
            }

        return {
            "covered": True,
        }

    return None


def get_fill_signature(cell):
    fill = cell.fill

    if fill.fill_type is None:
        return {
            "type": None,
        }

    if fill.fill_type == "gray125":
        return {
            "type": "gray125",
        }

    if fill.fill_type == "solid":
        color = fill.fgColor

        return {
            "type": "solid",
            "colorType": color.type,
            "theme": (
                color.theme
                if color.type == "theme"
                else None
            ),
            "rgb": (
                color.rgb
                if color.type == "rgb"
                else None
            ),
            "tint": color.tint,
        }

    return {
        "type": fill.fill_type,
    }


def get_row_tone(first_cell, is_header):
    """
    Translate the workbook's existing alternating
    schedule colors into semantic row tones.

    Current workbook convention:
      theme 9 = green
      theme 7 = blue
    """
    if is_header:
        return "neutral"

    fill = first_cell.fill

    if fill.fill_type is None:
        return "neutral"

    if fill.fill_type != "solid":
        raise ScheduleAnomaly(
            "UNKNOWN_ROW_FILL",
            (
                f"Unexpected row fill pattern "
                f"'{fill.fill_type}' at "
                f"{first_cell.parent.title}!"
                f"{first_cell.coordinate}."
            ),
        )

    color = fill.fgColor

    if color.type != "theme":
        raise ScheduleAnomaly(
            "UNKNOWN_ROW_COLOR",
            (
                f"Unexpected row color type "
                f"'{color.type}' at "
                f"{first_cell.parent.title}!"
                f"{first_cell.coordinate}."
            ),
        )

    if color.theme == 9:
        return "green"

    if color.theme == 7:
        return "blue"

    raise ScheduleAnomaly(
        "UNKNOWN_ROW_COLOR",
        (
            f"Unexpected Excel theme color "
            f"{color.theme} at "
            f"{first_cell.parent.title}!"
            f"{first_cell.coordinate}."
        ),
    )


def is_unavailable_cell(cell):
    return cell.fill.fill_type == "gray125"


def classify_cell(
    cell,
    text,
    relative_column,
    is_header,
):
    """
    Assign a renderer-safe semantic role.

    Relative columns:
      0 = day/date
      1 = time
      2-6 = apparatus/client slots
      7 = instructor
    """
    clean_text = text.strip()
    upper_text = clean_text.upper()

    if is_header:
        return "column-header"

    if is_unavailable_cell(cell):
        return "unavailable"

    if upper_text == "OPEN":
        return "open"

    if upper_text.startswith("NO CLASSES -"):
        return "no-classes-block"

    if upper_text.startswith("CLASS CANCELED"):
        return "class-canceled"

    if upper_text.startswith("NO CLASS"):
        return "no-class"

    if upper_text.startswith("MAT:"):
        return "mat"

    if relative_column == 0:
        return "day"

    if relative_column == 1:
        return "time"

    if relative_column == 7:
        return "instructor"

    if 2 <= relative_column <= 6:
        return "client"

    raise ScheduleAnomaly(
        "UNKNOWN_CELL_ROLE",
        (
            f"Could not classify "
            f"{cell.parent.title}!"
            f"{cell.coordinate}."
        ),
    )


def validate_cell_fill(
    cell,
    role,
    row_tone,
):
    """
    Detect formatting that our renderer would otherwise
    silently ignore or flatten.
    """
    signature = get_fill_signature(cell)
    fill_type = signature["type"]

    if role == "unavailable":
        if fill_type != "gray125":
            raise ScheduleAnomaly(
                "UNAVAILABLE_FILL_MISMATCH",
                (
                    f"Unavailable cell "
                    f"{cell.parent.title}!"
                    f"{cell.coordinate} no longer "
                    "uses the expected hatch pattern."
                ),
            )

        return

    if fill_type not in {
        None,
        "solid",
    }:
        raise ScheduleAnomaly(
            "UNKNOWN_CELL_FILL",
            (
                f"Unexpected fill pattern "
                f"'{fill_type}' at "
                f"{cell.parent.title}!"
                f"{cell.coordinate}."
            ),
        )

    if row_tone == "neutral":
        return

    if fill_type != "solid":
        raise ScheduleAnomaly(
            "ROW_FILL_MISSING",
            (
                f"{cell.parent.title}!"
                f"{cell.coordinate} does not match "
                f"the expected {row_tone} row fill."
            ),
        )

    if signature["colorType"] != "theme":
        raise ScheduleAnomaly(
            "UNKNOWN_CELL_COLOR",
            (
                f"Unexpected color type at "
                f"{cell.parent.title}!"
                f"{cell.coordinate}."
            ),
        )

    expected_theme = (
        9
        if row_tone == "green"
        else 7
    )

    if signature["theme"] != expected_theme:
        raise ScheduleAnomaly(
            "ROW_COLOR_MISMATCH",
            (
                f"{cell.parent.title}!"
                f"{cell.coordinate} does not match "
                f"the expected {row_tone} row color."
            ),
        )


def validate_merge(
    cell,
    merge_info,
    relative_column,
    column_count,
):
    if not merge_info:
        return

    if merge_info.get("covered"):
        return

    rowspan = merge_info["rowspan"]
    colspan = merge_info["colspan"]

    if rowspan < 1 or colspan < 1:
        raise ScheduleAnomaly(
            "INVALID_MERGE",
            (
                f"Invalid merged range beginning at "
                f"{cell.parent.title}!"
                f"{cell.coordinate}."
            ),
        )

    if relative_column + colspan > column_count:
        raise ScheduleAnomaly(
            "MERGE_OUTSIDE_SCHEDULE",
            (
                f"Merged range beginning at "
                f"{cell.parent.title}!"
                f"{cell.coordinate} extends outside "
                "the public schedule range."
            ),
        )


def extract_geometry(worksheet, cells):
    """
    Preserve fixed Excel geometry once per week.
    """
    first_row = cells[0]

    column_widths = []

    for cell in first_row:
        width = worksheet.column_dimensions[
            cell.column_letter
        ].width

        column_widths.append(width)

    row_heights = []

    for excel_row in cells:
        row_number = excel_row[0].row

        height = worksheet.row_dimensions[
            row_number
        ].height

        row_heights.append(height)

    return {
        "columnWidths": column_widths,
        "rowHeights": row_heights,
    }


def validate_header(week):
    header_cells = week["rows"][0]["cells"]

    if len(header_cells) != EXPECTED_COLUMN_COUNT:
        raise ScheduleAnomaly(
            "HEADER_COLUMN_COUNT",
            (
                f"{week['key']} header contains "
                f"{len(header_cells)} cells; "
                f"expected {EXPECTED_COLUMN_COUNT}."
            ),
        )

    actual_headers = [
        cell["text"]
        for cell in header_cells
    ]

    if actual_headers != EXPECTED_HEADERS:
        raise ScheduleAnomaly(
            "UNEXPECTED_HEADER",
            (
                f"{week['key']} has an unexpected "
                f"header structure: {actual_headers}"
            ),
        )


def validate_week(week):
    if week["columnCount"] != EXPECTED_COLUMN_COUNT:
        raise ScheduleAnomaly(
            "COLUMN_COUNT",
            (
                f"{week['key']} contains "
                f"{week['columnCount']} columns; "
                f"expected {EXPECTED_COLUMN_COUNT}."
            ),
        )

    if week["rowCount"] < 2:
        raise ScheduleAnomaly(
            "TOO_FEW_ROWS",
            (
                f"{week['key']} contains "
                "too few schedule rows."
            ),
        )

    if not week["rows"]:
        raise ScheduleAnomaly(
            "NO_ROWS",
            (
                f"{week['key']} contains "
                "no schedule data."
            ),
        )

    validate_header(week)


def extract_week(workbook, web_name):
    resolved = resolve_defined_name(
        workbook,
        web_name,
    )

    worksheet = workbook[
        resolved["sheet"]
    ]

    cells = worksheet[
        resolved["range"]
    ]

    column_count = len(cells[0])

    rows = []

    for row_index, excel_row in enumerate(cells):
        is_header = row_index == 0

        row_tone = get_row_tone(
            excel_row[0],
            is_header,
        )

        row_data = {
            "tone": row_tone,
            "cells": [],
        }

        for relative_column, cell in enumerate(excel_row):
            merge_info = find_merge_for_cell(
                worksheet,
                cell,
            )

            if (
                merge_info
                and merge_info.get("covered")
            ):
                continue

            validate_merge(
                cell,
                merge_info,
                relative_column,
                column_count,
            )

            value = cell.value

            if value is None:
                text = ""
            else:
                text = str(value)

            role = classify_cell(
                cell,
                text,
                relative_column,
                is_header,
            )

            validate_cell_fill(
                cell,
                role,
                row_tone,
            )

            cell_data = {
                "text": text,
                "style": role,
            }

            if cell.font.bold:
                cell_data["bold"] = True

            if merge_info:
                if merge_info["rowspan"] > 1:
                    cell_data["rowspan"] = (
                        merge_info["rowspan"]
                    )

                if merge_info["colspan"] > 1:
                    cell_data["colspan"] = (
                        merge_info["colspan"]
                    )

            row_data["cells"].append(
                cell_data
            )

        rows.append(row_data)

    geometry = extract_geometry(
        worksheet,
        cells,
    )

    week = {
        "key": web_name,
        "sourceName": resolved.get(
            "points_to",
            web_name,
        ),
        "sourceSheet": resolved["sheet"],
        "sourceRange": resolved["range"],
        "rowCount": len(cells),
        "columnCount": column_count,
        "columnWidths":
            geometry["columnWidths"],
        "rowHeights":
            geometry["rowHeights"],
        "rows": rows,
    }

    validate_week(week)

    return week


def write_json_atomically(
    output_path,
    schedule_data,
):
    """
    Write a complete candidate first.

    The existing known-good schedule.json is replaced
    only after extraction and all validation succeed.
    """
    temp_path = output_path.with_suffix(
        ".json.tmp"
    )

    with temp_path.open(
        "w",
        encoding="utf-8",
    ) as output_file:
        json.dump(
            schedule_data,
            output_file,
            indent=2,
            ensure_ascii=False,
        )

        output_file.write("\n")

    temp_path.replace(
        output_path
    )


def main():
    workbook_path = get_workbook_path()
    output_path = Path(DEFAULT_OUTPUT)

    info(
        "EXTRACT_START",
        f"Opening workbook: {workbook_path}",
    )

    if not workbook_path.exists():
        raise FileNotFoundError(
            f"Workbook does not exist: "
            f"{workbook_path}"
        )

    workbook = load_workbook(
        workbook_path,
        data_only=True,
        read_only=False,
    )

    web_names = [
        name
        for name in workbook.defined_names.keys()
        if name.startswith("Web_")
    ]

    if not web_names:
        raise ScheduleAnomaly(
            "NO_WEB_ALIASES",
            (
                "No workbook-level defined names "
                "beginning with 'Web_' were found."
            ),
        )

    for web_name in web_names:
        validate_web_name(
            web_name
        )

    web_names = sorted(
        web_names,
        key=web_name_sort_key,
    )

    info(
        "WEB_ALIASES_FOUND",
        (
            f"Found {len(web_names)} "
            "web schedule aliases."
        ),
    )

    weeks = []

    for web_name in web_names:
        resolved = resolve_defined_name(
            workbook,
            web_name,
        )

        intermediate = resolved.get(
            "points_to"
        )

        if intermediate:
            info(
                "RANGE_RESOLVED",
                (
                    f"{web_name}"
                    f" -> {intermediate}"
                    f" -> {resolved['sheet']}"
                    f"!{resolved['range']}"
                ),
            )
        else:
            info(
                "RANGE_RESOLVED",
                (
                    f"{web_name}"
                    f" -> {resolved['sheet']}"
                    f"!{resolved['range']}"
                ),
            )

        weeks.append(
            extract_week(
                workbook,
                web_name,
            )
        )

    schedule_data = {
        "schemaVersion": 2,
        "weeks": weeks,
    }

    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    write_json_atomically(
        output_path,
        schedule_data,
    )

    info(
        "JSON_CREATED",
        f"Created: {output_path}",
    )

    info(
        "EXTRACT_COMPLETE",
        (
            f"Extracted {len(weeks)} "
            "web schedule weeks."
        ),
    )

    print(
        "SUCCESS: Public schedule JSON "
        "passed validation and was generated."
    )


if __name__ == "__main__":
    try:
        main()

    except ScheduleAnomaly as error:
        print()
        print(
            "::error::"
            f"LEVEL 2 SCHEDULE ANOMALY "
            f"[{error.code}]: "
            f"{error.message}"
        )
        print(
            "Schedule JSON was NOT published."
        )
        print(
            "The last known-good live schedule "
            "remains unchanged."
        )

        sys.exit(2)

    except Exception as error:
        print()
        print(
            "::error::"
            "LEVEL 1 PIPELINE FAILURE "
            "[EXTRACTOR_FAILURE]: "
            f"{type(error).__name__}: {error}"
        )
        print(
            "Schedule JSON was NOT published."
        )
        print(
            "The last known-good live schedule "
            "remains unchanged."
        )

        sys.exit(1)